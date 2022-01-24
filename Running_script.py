# -*- coding: utf-8 -*-
"""
Created on Mon May 25 16:07:50 2020

@author: Vasco Saltão

Script that runs the tool


LIST OF SYMBOLS AND ACRONYMS:
    
    r - Initial rate
    P - Total pressure
    Pi - Partial pressure of i
    yi - Molar fraction of i
    Ki - Equilibrium constant of adsorption of i
    k - Kinetic constant of reaction
    ID - Identification of mechanism
    MSE - Mean squared error
    SRE - Sum of relative errors
"""

import os
from pathlib import Path
import matplotlib.pyplot as plt
import Library_module as lib
import FeatureExtraction as ft
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


"""
Obtaining the experimental dataset
"""
Tool_path = Path().absolute()

while True:
    
    try:
        
        filename = str(input("Please type name of Excel file inserted in the 'Input' folder(with the extension included)\n\n"))
        
        file_path = Tool_path / "Input" / filename

        dataset = pd.read_excel(file_path) # Excel file that the user must fill with the experimental dataset
        
        break
        
    except FileNotFoundError:
        
        print("Please input the correct name of the file in the 'Input' folder")


pressure = list(dataset['P'])

Exp_rate = list(dataset['r0'])

good_dataset = True

if (len(pressure) != len(Exp_rate)) or (len(dataset) == 0):
    
    good_dataset = False
    print("Sorry, the dataset in the Excel file 'Dataset_Template.xlsx' is not valid. Please insert valid dataset and try again")
    
else:
    
    print('Pressure values:', pressure)
    
    P_units = str(input('Please input the pressure units of these values:   '))
    
    print('Initial rate values:', Exp_rate)
    
    r0_units = str(input('Please input the rate units of these values:   '))

"""
Asking user for input
"""
if good_dataset == True:
    
    R_names = []
    R_stoich = [0, 0]
    R_frac = [0, 0]
    
    P_names = []
    P_stoich = [0, 0]
    
    while True:
        
        try:
            
            number_of_react = int(input('How many reactants?   '))
            
            good_react_input = True
            break
            
        except ValueError:
            
            print('\n', 'Please insert an integer number of reactants')
                
            
    if number_of_react > 2:
                
        good_react_input = False
        
        print('Sorry, this tool does not support reactions with more than 2 reactants at the moment')
        
    else:
        
        i=0
        while i < number_of_react:
            
            R_names.append(str(input('Name of reactant {}:   '.format(i+1))))
            
            while True:
                
                try:
                    
                    R_stoich[i] = float(input('Stoichiometric coefficient of reactant {} in reaction:   '.format(i+1)))
                    
                    break
                
                except ValueError:
                    
                    print('\n', 'Please insert a valid input')
                    
            while True:
                
                try:
                    
                    R_frac[i] = float(input('Molar fraction of reactant {} in feed (between 0 and 1):   '.format(i+1)))
                    
                    if R_frac[i] < 0 or R_frac[i] > 1:
                        
                        print('\n', 'Please insert a valid input (between 0 and 1)')
                        
                    else:
                        
                        break
                
                except ValueError:
                    
                    print('\n', 'Please insert a valid input')
                    
            i+=1
        
    
if good_dataset == True and good_react_input == True:
    
    while True:
            
        try:
            
            number_of_prod = int(input('How many products?   '))
            
            good_prod_input = True
            break
        
        except ValueError:
            
            print('\n', 'Please insert an integer number of products')
    
    if number_of_prod > 2:
        
        good_prod_input = False
        
        print('Sorry, this tool does not support reactions with more than 2 products at the moment')
        
    else:
        
        j=0
        while j < number_of_prod:
            
            P_names.append(str(input('Name of product {}:   '.format(j+1))))
            
            while True:
            
                try:
                    
                    P_stoich[j] = int(input('Stoichiometric number of product {} in reaction:   '.format(j+1)))
                    
                    break
                
                except ValueError:
                    
                    print('\n', 'Please insert a valid input')
            
            j+=1


if good_dataset == True and good_react_input == True and good_prod_input == True:
    
    while True:
               
        try:
            
            yI = float(input('Molar fraction of inerts in feed (between 0 and 1, 0 if none):   '))
            
            if yI < 0 or yI > 1:
                        
                print('\n', 'Please insert a valid input (between 0 and 1)')
                
            else:
                
                break
        
        except ValueError:
                
            print('\n', 'Please insert a valid input')
    
    good_inert_input = True
    
    if sum(R_frac) + yI != 1:
        
        good_inert_input = False
        
        print('ERROR: Molar fractions of reactants and inert species does not add up to 1')

    
if good_dataset == True and good_react_input == True and good_prod_input == True and good_inert_input == True:
    
    Output_name = str(input('Please type the desired name for the folder with the output:\n\n'))
    
    Output_path = Tool_path / "Output" / Output_name
    
    os.mkdir(Output_path)
    
    
    """
    Extracting the experimental features
    """
    Exp_rate_norm = [x/max(Exp_rate) for x in Exp_rate]
    
    plt.figure(num = 18)
    
    Exp_Features = ft.features(pressure, Exp_rate_norm)
    
    Exp_Primitives = Exp_Features[0]
    Exp_Extremes = Exp_Features[1]
    
    max_rate = max(Exp_rate_norm)
    max_y_values = max(Exp_Features[2])
    
    plt.ylim(0, 1.2*max(max_y_values, max_rate))
    
    plt.title('Experimental curve features')
        
    plt.xlabel('P (' + P_units + ')')
    plt.ylabel('r0 / max(r0) ')
    
    plt.savefig(Output_path / 'Experimental features')
    
    """
    User eliminates mechs 'a priori' for A = R1 and B = R2
    """
    a = R_stoich[0]
    b = R_stoich[1]
    r = P_stoich[0]
    s = P_stoich[1]
    
    print('\n\n\n')
    
    for ID in lib.Tags_Lib:
        
        Tags = lib.Tags_Lib[ID]
        
        if b == 0 and ('Adsorption of B controlling' in Tags or 'B does not adsorb' in Tags or 'A does not adsorb' in Tags):
            
            continue
        
        Output_tags = []
        
        for Tag in Tags:
        
            New_Tag = Tag
            
            New_Tag = New_Tag.replace('A ', R_names[0] + ' ')
            
            if b > 0:
                
                New_Tag = New_Tag.replace('B ', R_names[1] + ' ')
            
            if s == 0:
                
                New_Tag = New_Tag.replace('R ', P_names[0] + ' ')
            
            else:
                
                New_Tag = New_Tag.replace('R ', 'a product ')
            
            Output_tags.append(New_Tag)
        
        if ID == 'r':
            
            Output_tags = ['Uncatalyzed reaction']
        
        print(ID, ':', Output_tags)
    
    Bad_IDs_1_str = list(input('If you know that any of the mechanisms cannot be the corect one, please write them below, separated by commas. If no, leave it blank and just press enter.\n\n'))
    
    if len(Bad_IDs_1_str) == 0:
        
        Bad_IDs_1 = []
    
    else:
        
        Bad_IDs_1 = [x.strip() for x in Bad_IDs_1_str.split(',')]
    
    if number_of_react > 1:
        
        """
        User eliminates mechs 'a priori' for A = R2 and B = R1
        """
        a = R_stoich[1]
        b = R_stoich[0]
        
        print('\n\n\n')
            
        for ID in lib.IDs_swap:
            
            Tags = lib.Tags_Lib[ID]
            
            Output_tags = []
        
            for Tag in Tags:
                
                New_Tag = Tag
                
                New_Tag = New_Tag.replace('A ', R_names[1] + ' ')
            
                New_Tag = New_Tag.replace('B ', R_names[0] + ' ')
            
                if s == 0:
                    
                    New_Tag = New_Tag.replace('R ', P_names[0] + ' ')
                
                else:
                    
                    New_Tag = New_Tag.replace('R ', 'a product ')
                    
                Output_tags.append(New_Tag)
        
            print(ID, ':', Output_tags)
    
        Bad_IDs_2_str = list(input('Now please do the same to the mechanisms above (same mechanisms, but the reactants are swapped).\n\n'))
        
        if len(Bad_IDs_2_str) == 0:
        
            Bad_IDs_2 = []
            
        else:
            
            Bad_IDs_2 = [x.strip() for x in Bad_IDs_2_str.split(',')]
    
    """
    Asking the user if they want to use the smaller list of values to save 
    computational time (if necessary)
    """
    Redux = False
    
    if (number_of_react > 1) and (number_of_prod == 1):
        
        if Bad_IDs_1 == ['c', 'g', 'j', 'm', 'p'] and Bad_IDs_2 == ['g', 'j', 'm', 'p']:
            
            pass
        
        else:
            
            correct_input = False
                
            while correct_input == False:
                
                Answer_Redux = str(input('This specific case (2 reactants, 1 product) can lead to extensive runtimes (aprox. 4h) if the full list of K values is used to build theoretical curves for mechanisms where the RDS is the desorption of a product (c, g, j, m and p). \n\nDo you wish to use the smaller list of values to reduce the runtime? (Only the previously mentioned mechanisms will be affected by this decision) (Y/N)   '))
                
                if Answer_Redux in ['N', 'n', 'No', 'no']:
                    
                    correct_input = True
                    
                elif Answer_Redux in ['Y', 'y', 'Yes', 'yes']:
                        
                    correct_input = True
                    Redux = True
                    
                else:
                    
                    print('Please insert correct input')
                    
    
    """
    A = R1, B = R2 - Building the theoretical curves
    """
    R_names_1 = R_names
    
    a = R_stoich[0]
    b = R_stoich[1]
    
    Stoich_1 = [a, b] + P_stoich # [a, b, r, s] in a reaction aA + bB <--> rR + sS

    yA_1 = R_frac[0]
    yB_1 = R_frac[1]
    
    plt.figure(num = 19)
    all_Theor_Features_1 = lib.Theoretical_Features(Stoich_1, yA_1, yB_1, pressure, Bad_IDs_1, Redux)
    plt.close(19)
    
    if number_of_react > 1:
        
        """
        A = R2, B = R1 - Building the theoretical curves
        """
        R_names_2 = [R_names[1], R_names[0]]
        
        a = R_stoich[1]
        b = R_stoich[0]
        
        Stoich_2 = [a, b] + P_stoich # [a, b, r, s] in a reaction aA + bB <--> rR + sS
        
        yA_2 = R_frac[1]
        yB_2 = R_frac[0]
        
        plt.figure(num = 20)
        all_Theor_Features_2 = lib.Theoretical_Features(Stoich_2, yA_2, yB_2, pressure, Bad_IDs_2, Redux, Swap_run=True)
        plt.close(20)
    
    """
    Comparing experimental and theoretical features
    """
    plt.show(block=False) # Plotting the graph with the experimental features
    
    print('\n\n', 'Experimental primitives:', Exp_Primitives)
    print('\n', 'Pressure values of experimental extremes ({}):'.format(P_units), Exp_Extremes)
    
    print('\n\n', 'Now the simulated datasets will be compared with the experimental data. This will be made in the following manner:', '\n')
    
    print('\n', '   - Check if both curves have the same set of primitives;')
    print('\n', '   - If so, check if the value of the extremes (borders between primitives) of the theoretical curve are the same as the extremes of the experimental curve, within a certain range of tolerance (which is a user input)')
    
    run = True
    while run == True:
        
        Tolerance = float(input('Please input tolerance value (between 0 and 1) for comparison of simulated datasets with experimental data:   '))
        
        
        Tables_1 = lib.Lib_Check(all_Theor_Features_1, Exp_Features, Tolerance, Exp_rate_norm)
        
        Final_table_1, Text_list_1, Tags_list_1, Theor_press = lib.Exp_and_Theoretical(Stoich_1, yA_1, yB_1, pressure, Exp_rate, Tables_1, R_names_1, P_names)
        
        
        if number_of_react > 1:
            
            Tables_2 = lib.Lib_Check(all_Theor_Features_2, Exp_Features, Tolerance, Exp_rate_norm)
            
            Final_table_2, Text_list_2, Tags_list_2, Theor_press = lib.Exp_and_Theoretical(Stoich_2, yA_2, yB_2, pressure, Exp_rate, Tables_2, R_names_2, P_names, Swap=True)
            
            
            Text_list = dict(Text_list_1, **Text_list_2)
            
            Tags_list = dict(Tags_list_1, **Tags_list_2)
            
            Final_table = pd.concat([Final_table_1, Final_table_2], ignore_index=True)
        
        else:
            
            Text_list = Text_list_1
            
            Tags_list = Tags_list_1
            
            Final_table = Final_table_1
        
        Final_table = Final_table.sort_values(by = ['SRE', 'MSE'], ignore_index = True)
        
        Final_table = Final_table.replace({'KA': 0, 'KB': 0, 'KR': 0, 'KS': 0, 'Kr': 0}, 'N/A')
        
        lib.Create_output(Final_table, Text_list, Tags_list, Theor_press, pressure, Exp_rate, P_units, r0_units)
        
        if number_of_react > 1:
            
            print('\n\n', 'Mechanisms with _1 : A = {}, B = {}'.format(R_names[0], R_names[1]))
            
            print('\n', 'Mechanisms with _2 : A = {}, B = {}'.format(R_names[1], R_names[0]))
        
        
        correct_input = False
        
        while correct_input == False:
            
            retry = str(input('Try again with different tolerance?\n\nATTENTION: The tool will only save the output from the last run\n\n(Y/N):   '))
            
            if retry in ['N', 'n', 'No', 'no']:
                
                correct_input = True
                run = False
                
            elif retry in ['Y', 'y', 'Yes', 'yes']:
                
                correct_input = True
                
            else:
                
                print('Please insert correct input')
    
    
    """
    Saving the output
    """
    Output_table = Final_table.drop('Theor_rate', 1)
    
    wb = Workbook()
    
    ws = wb.active
    
    ws['A1'] = 'Exp data from:'
    ws['B1'] = filename
    
    ws['A3'] = 'Primitives:'
    
    i=0
    while i < len(Exp_Primitives):
        
        ws.cell(row=3, column=i+2, value=Exp_Primitives[i])
        
        i+=1
    
    ws['A4'] = 'Extremes:'
    
    i=0
    while i < len(Exp_Extremes):
        
        ws.cell(row=4, column=i+2, value=Exp_Extremes[i])
        
        i+=1
    
    ws['A6'] = 'Tolerance:'
    ws['B6'] = Tolerance
    
    ws['A8'] = 'Proposed mechanisms (in order of best match):'
    
    rows = dataframe_to_rows(Output_table, index=False, header=True)
    
    for r_idx, row in enumerate(rows, 1):
        
        for c_idx, value in enumerate(row, 1):
            
            ws.cell(row=r_idx+10, column=c_idx, value=value)
    
    i=0
    while i < len(Final_table):
        
        ID = Final_table['ID'][i]
        
        Tags = Tags_list[ID]
        
        Theor_rate = Final_table['Theor_rate'][i]
        
        plt.figure(i+1)
        
        Title = 'Mechanism ' + str(ID) + ": " + str(Tags)
        
        plt.title(Title)
        
        plt.xlabel('P (' + P_units + ')')
        plt.ylabel('r0 (' + r0_units + ')')
        
        plt.plot(pressure, Exp_rate, 'ro', label='Experimental') # plotting the experimental datapoints
        
        plt.plot(Theor_press, Theor_rate, color='tab:blue', label='Theoretical')
        
        plt.legend()
        
        plt.savefig(Output_path / ID)
        
        plt.close(i+1)
        
        text = Text_list[ID]
        
        rows = text.split('\n')
        
        ws.cell(row=1, column=i+12, value=ID)
        
        j=0
        while j < len(rows):
            
            ws.cell(row=j+2, column=i+12, value=rows[j])
            
            j+=1
        
        i+=1
    
    wb.save(Output_path / "Proposed mechanisms.xlsx")